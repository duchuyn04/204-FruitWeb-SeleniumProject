"""
=============================================================================
auto_format_testcases.py
=============================================================================
🎯 MỤC ĐÍCH:
    Tự động format (làm đẹp) sheet test case trong file Excel 204_local.xlsx.
    Script sẽ kẻ viền, căn lề, in đậm, tô màu và gộp ô (merge) đúng chuẩn
    cho toàn bộ các khối Test Case trong sheet được chỉ định.

📋 NHỮNG GÌ SCRIPT LÀM:
    - Kẻ viền (border) cho tất cả các ô trong vùng dữ liệu
    - In đậm các cột quan trọng: No, Req ID, TC ID, Step#, Expected  
    - Căn giữa theo chiều dọc/ngang cho các cột tiêu đề
    - Căn trái cho cột F, G, H (Step#, Step Action, Test Data)
    - Tô màu vàng cột I (Expected Result) để nổi bật
    - Gộp ô (Merge) cột A, B, C, D, E, I, J, K, L, M theo từng Test Case

⚙️ CÁCH SỬ DỤNG:
    1. Đóng file 204_local.xlsx trong Excel (BẮT BUỘC!)
    2. Thay đổi SHEET_NAME bên dưới nếu muốn format sheet khác
    3. Mở Terminal, chạy: python auto_format_testcases.py

📌 BẢNG THAM KHẢO SHEET_NAME & START_ROW:
    ┌─────────────────────────────┬──────────────────────────┬───────────┐
    │ Module                      │ SHEET_NAME               │ START_ROW │
    ├─────────────────────────────┼──────────────────────────┼───────────┤
    │ Order Management            │ TC_OrderManagement       │ 12        │
    │ Product Management          │ TC_Product Management    │ 12        │
    │ Authentication              │ TC_Authentication        │ 12        │
    │ Cart Management             │ TC_Cart Management       │ 12        │
    │ Category Management         │ TC_Category management   │ 12        │
    └─────────────────────────────┴──────────────────────────┴───────────┘

⚠️ LƯU Ý:
    - Script nhận diện đầu Test Case dựa vào cột A (No.) có giá trị
    - Chạy lại script sau mỗi lần cập nhật/sắp xếp dữ liệu trong sheet
=============================================================================
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==========================================
# ⚙️ CẤU HÌNH BAN ĐẦU (TÙY CHỈNH TẠI ĐÂY)
# ==========================================
# Tự động tính đường dẫn tuyệt đối dựa theo vị trí script (không cần sửa)
_SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT   = os.path.dirname(_SCRIPT_DIR)
EXCEL_FILE_PATH = os.path.join(_PROJECT_ROOT, '204_local.xlsx')

SHEET_NAME      = 'TC_OrderManagement' # Tên sheet cần format (xem bảng trên)
START_ROW       = 12                   # Hàng dữ liệu đầu tiên (sau dòng Header)
ROW_HEIGHT      = 65                   # Chiều cao mỗi dòng (đơn vị: point, mặc định 60)
                                       # Gợi ý: 40=thấp, 60=vừa, 80=cao, None=tự động
# ==========================================


def format_test_cases():
    print(f"[*] Đang tải load file Excel: {EXCEL_FILE_PATH}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb[SHEET_NAME]
    except Exception as e:
        print(f"[!] Lỗi không thể mở file hoặc tìm thấy sheet: {e}")
        return

    # -------------------------------------------------------------------------
    # 1. KHỞI TẠO CÁC MẪU ĐỊNH DẠNG (FORMAT TEMPLATES)
    #    Tạo sẵn các mẫu định dạng để tái sử dụng, tránh tạo lặp lại
    # -------------------------------------------------------------------------
    font_bold = Font(name='Arial', size=11, bold=True)   # Font in đậm
    font_norm = Font(name='Arial', size=11, bold=False)  # Font thường

    # Căn giữa cả ngang+dọc, tự động xuống dòng khi nội dung dài
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # Căn trái, tự động xuống dòng (dùng cho Step Action, Test Data)
    align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Màu nền vàng cho cột Expected Result (giúp QA nhìn nhanh kết quả mong đợi)
    fill_yellow = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    # Kiểu viền mỏng, áp dụng cho tất cả 4 cạnh của ô
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # -------------------------------------------------------------------------
    # 2. HÀM ÁP DỤNG ĐỊNH DẠNG CHO 1 KHỐI (1 TEST CASE)
    #    Mỗi Test Case có thể gồm nhiều dòng (do có nhiều Step).
    #    Hàm này nhận vào dòng đầu (start_r) và dòng cuối (end_r) của khối đó,
    #    rồi tự động merge + format toàn bộ vùng đó.
    # -------------------------------------------------------------------------
    def apply_block_format(start_r, end_r):
        # Các cột cần Merge theo chiều dọc trong 1 Test Case:
        #   A=No, B=Req ID, C=TC ID, D=Objective, E=Preconditions,
        #   I=Expected, J=Actual, K=Screenshot, L=Result, M=Notes
        cols_to_merge = [1, 2, 3, 4, 5, 9, 10, 11, 12, 13]

        # Chỉ merge khi Test Case có từ 2 dòng trở lên (nhiều Step)
        if end_r > start_r:
            for c in cols_to_merge:
                ws.merge_cells(
                    start_row=start_r, start_column=c,
                    end_row=end_r,     end_column=c
                )

        # Áp dụng định dạng cho từng ô trong khối
        for r in range(start_r, end_r + 1):
            # Áp dụng chiều cao hàng (set 1 lần cho mỗi row, ngoài vòng for c)
            if ROW_HEIGHT is not None:
                ws.row_dimensions[r].height = ROW_HEIGHT

            for c in range(1, 14):  # Cột A (1) đến M (13)
                cell = ws.cell(row=r, column=c)

                # Kẻ viền cho tất cả ô
                cell.border = thin_border

                # In đậm cho: No(A), Req ID(B), TC ID(C), Step#(F), Expected(I), Result(L)
                if c in [1, 2, 3, 6, 9, 12]:
                    cell.font = font_bold
                else:
                    cell.font = font_norm

                # Căn trái cho cột F=Step#, G=Step Action, H=Test Data
                # Căn giữa cho tất cả các cột còn lại
                if c in [6, 7, 8]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center

                # Tô nền vàng cho cột I (Expected Result) để nổi bật khi review
                if c == 9:
                    cell.fill = fill_yellow

    # -------------------------------------------------------------------------
    # 3. QUÉT TOÀN BỘ SHEET VÀ NHẬN DIỆN CÁC KHỐI TEST CASE
    #    Logic: Khi cột A (No.) có giá trị → đó là dòng đầu của Test Case mới
    #           → Format khối Test Case trước đó, rồi bắt đầu khối mới
    # -------------------------------------------------------------------------
    print(f"[*] Đang xử lý data bắt đầu từ dòng {START_ROW}...")
    max_r = ws.max_row
    current_block_start = START_ROW  # Mốc theo dõi dòng đầu của khối hiện tại

    for r in range(START_ROW + 1, max_r + 1):
        cell_val = ws.cell(row=r, column=1).value  # Đọc giá trị cột A

        # Nếu cột A có giá trị → bắt đầu 1 Test Case mới
        if cell_val is not None and str(cell_val).strip() != '':
            # Format khối Test Case vừa kết thúc (từ current_block_start đến r-1)
            apply_block_format(current_block_start, r - 1)
            # Cập nhật lại mốc bắt đầu của khối mới
            current_block_start = r

    # Xử lý khối Test Case cuối cùng (vòng lặp trên không format được khối cuối)
    apply_block_format(current_block_start, max_r)

    # -------------------------------------------------------------------------
    # 4. LƯU LẠI KẾT QUẢ
    # -------------------------------------------------------------------------
    wb.save(EXCEL_FILE_PATH)
    print("[v] Hoàn tất! File Excel đã được format đẹp mắt 100%.")


if __name__ == "__main__":
    format_test_cases()
