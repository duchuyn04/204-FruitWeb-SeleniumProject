"""
=============================================================================
import_md_to_excel.py
=============================================================================
🎯 MỤC ĐÍCH:
    Đọc dữ liệu thô từ file Markdown (.md) chứa bảng Test Case,
    sau đó ghi vào đúng sheet trong file Excel 204_local.xlsx.
    Sau khi ghi xong, tự động chạy luôn auto_format_testcases.py để làm đẹp.

📋 NHỮNG GÌ SCRIPT LÀM:
    1. Đọc và parse bảng test case từ file .md
    2. Sắp xếp theo thứ tự Requirement ID (F12.1 → F12.2 → F12.3 ...)
    3. Unmerge toàn bộ ô đã gộp cũ trong sheet
    4. Xóa dữ liệu cũ cột A-I (giữ nguyên kết quả test J-M)
    5. Ghi lại dữ liệu mới đúng thứ tự, merge ô theo từng khối Test Case
    6. Lưu file Excel

⚙️ CÁCH SỬ DỤNG:
    1. Đóng file 204_local.xlsx trong Excel (BẮT BUỘC!)
    2. Thay đổi MD_PATH và SHEET_NAME bên dưới nếu cần
    3. Mở Terminal, chạy:
           python import_md_to_excel.py
    4. Sau đó chạy tiếp để format cho đẹp:
           python auto_format_testcases.py

📌 BẢNG THAM KHẢO CẤU HÌNH:
    ┌──────────────────────────────────────────────────────────────────┐
    │ Module           │ MD_PATH (tên file trong markdown/test case/)  │
    ├──────────────────┼───────────────────────────────────────────────┤
    │ Order Mgmt       │ test_cases_order_F12.1-F12.25.md             │
    │ Product Mgmt     │ (tên file .md tương ứng)                     │
    └──────────────────┴───────────────────────────────────────────────┘

⚠️ LƯU Ý:
    - Script giữ nguyên dữ liệu cột J, K, L, M (Actual, Screenshot, 
      Result, Notes) khi TC_ID khớp với dữ liệu cũ.
    - Script nhận diện dòng đầu của Test Case khi cột "No" có giá trị.
=============================================================================
"""

import re
import os
import codecs
import openpyxl

# ==========================================
# ⚙️ CẤU HÌNH BAN ĐẦU (TÙY CHỈNH TẠI ĐÂY)
# ==========================================
# Tự động tính đường dẫn tuyệt đối dựa theo vị trí script (không cần sửa)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)   # Lên 1 cấp = SeleniumProject/
_MD_BASE    = os.path.join(_PROJECT_ROOT, 'markdown', 'test case')

EXCEL_PATH     = os.path.join(_PROJECT_ROOT, '204_local.xlsx')
SHEET_NAME     = 'TC_OrderManagement'
MD_PATH        = os.path.join(_MD_BASE, 'test_cases_order_F12.1-F12.25.md')
DATA_START_ROW = 12   # Hàng bắt đầu ghi dữ liệu (sau dòng header)
# ==========================================


def sort_key(first_row):
    """Hàm tạo khóa sắp xếp theo Requirement ID và số thứ tự TC."""
    req_id = first_row[1].strip()   # Ví dụ: F12.13
    tc_id  = first_row[2].strip()   # Ví dụ: TC_F12.13_01
    m1 = re.search(r'F12\.(\d+)', req_id)
    m2 = re.search(r'_(\d+)$', tc_id)
    return (int(m1.group(1)) if m1 else 0,
            int(m2.group(1)) if m2 else 0)


def parse_markdown(md_path):
    """Đọc file Markdown và trả về danh sách các Test Case."""
    lines = codecs.open(md_path, encoding='utf-8').read().splitlines()

    table_lines = []
    in_table = False
    for line in lines:
        if line.startswith('| No |') or line.startswith('|:---'):
            in_table = True
            continue
        if in_table and line.startswith('|'):
            table_lines.append(line)

    def parse_cols(line):
        parts = line.split('|')[1:-1]
        return [col.strip() for col in parts]

    parsed_rows = []
    current_row = []
    for line in table_lines:
        if not line.strip():
            continue
        cols = parse_cols(line)
        if len(cols) < 9:
            continue
        if cols[0]:  # Cột No có giá trị → Bắt đầu Test Case mới
            if current_row:
                parsed_rows.append(current_row)
            current_row = [cols]
        else:        # Cột No trống → Dòng Step phụ của TC hiện tại
            current_row.append(cols)

    if current_row:
        parsed_rows.append(current_row)

    return parsed_rows


def import_md_to_excel():
    # BƯỚC 1: Đọc và parse markdown
    print(f"[*] Đọc dữ liệu từ: {MD_PATH}")
    parsed_rows = parse_markdown(MD_PATH)
    parsed_rows.sort(key=lambda r: sort_key(r[0]))
    print(f"    → Tìm thấy {len(parsed_rows)} Test Case")

    # BƯỚC 2: Mở file Excel
    print(f"[*] Mở file Excel: {EXCEL_PATH} | Sheet: {SHEET_NAME}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME]
    except Exception as e:
        print(f"[!] Lỗi mở file: {e}")
        print("    → Hãy đảm bảo file Excel đang ĐÓNG trước khi chạy script!")
        return

    max_row = ws.max_row

    # BƯỚC 3: Unmerge toàn bộ ô đã gộp trong vùng dữ liệu (row 11 trở đi)
    print(f"[*] Đang unmerge các ô đã gộp cũ...")
    merges_to_remove = [
        str(m) for m in ws.merged_cells.ranges if m.min_row >= DATA_START_ROW
    ]
    for ref in merges_to_remove:
        ws.unmerge_cells(ref)
    print(f"    → Đã unmerge {len(merges_to_remove)} vùng")

    # BƯỚC 4: Lưu kết quả test cũ (cột J, K, L, M) theo TC_ID
    old_results = {}
    for r in range(DATA_START_ROW, max_row + 1):
        tc_id = ws.cell(row=r, column=3).value
        if tc_id and isinstance(tc_id, str) and tc_id.startswith('TC_'):
            old_results[tc_id] = {
                'J': ws.cell(row=r, column=10).value,
                'K': ws.cell(row=r, column=11).value,
                'L': ws.cell(row=r, column=12).value,
                'M': ws.cell(row=r, column=13).value,
            }
    print(f"[*] Lưu lại kết quả test của {len(old_results)} TC cũ")

    # BƯỚC 5: Xóa dữ liệu cũ cột A-I (giữ nguyên J-M)
    for r in range(DATA_START_ROW, max_row + 1):
        for c in range(1, 10):
            ws.cell(row=r, column=c).value = None
    print(f"[*] Đã xóa dữ liệu cũ (A-I) từ row {DATA_START_ROW} đến {max_row}")

    # BƯỚC 6: Ghi dữ liệu mới từ markdown vào Excel
    print(f"[*] Đang ghi dữ liệu mới...")
    current_row_num = DATA_START_ROW
    for tc_index, tc_data in enumerate(parsed_rows):
        first_step = tc_data[0]
        no        = str(tc_index + 1)
        req_id    = first_step[1]
        tc_id     = first_step[2]
        objective = first_step[3]
        precond   = first_step[4]
        expected  = first_step[8]

        tc_start = current_row_num

        for step_idx, step_row in enumerate(tc_data):
            step_num    = step_row[5]
            step_action = step_row[6]
            test_data   = step_row[7]

            if step_idx == 0:  # Dòng đầu của Test Case: ghi đủ thông tin
                ws.cell(row=current_row_num, column=1).value = no
                ws.cell(row=current_row_num, column=2).value = req_id
                ws.cell(row=current_row_num, column=3).value = tc_id
                ws.cell(row=current_row_num, column=4).value = objective
                ws.cell(row=current_row_num, column=5).value = precond
                ws.cell(row=current_row_num, column=9).value = expected

            # Các dòng Step: chỉ ghi F, G, H
            ws.cell(row=current_row_num, column=6).value = step_num
            ws.cell(row=current_row_num, column=7).value = step_action
            ws.cell(row=current_row_num, column=8).value = test_data

            current_row_num += 1

        tc_end = current_row_num - 1

        # Merge ô nếu Test Case có nhiều Step
        if len(tc_data) > 1:
            for col in [1, 2, 3, 4, 5, 9, 10, 11, 12, 13]:
                ws.merge_cells(
                    start_row=tc_start, start_column=col,
                    end_row=tc_end,     end_column=col
                )

        # Khôi phục kết quả test cũ nếu TC_ID khớp
        if tc_id in old_results:
            res = old_results[tc_id]
            ws.cell(row=tc_start, column=10).value = res['J']
            ws.cell(row=tc_start, column=11).value = res['K']
            ws.cell(row=tc_start, column=12).value = res['L']
            ws.cell(row=tc_start, column=13).value = res['M']

    # BƯỚC 7: Lưu file
    wb.save(EXCEL_PATH)
    print(f"[v] Ghi xong! Tổng {len(parsed_rows)} Test Case đến row {current_row_num - 1}")
    print()
    print("=" * 60)
    print("  Bước tiếp theo: Chạy lệnh sau để format đẹp Excel:")
    print("      python auto_format_testcases.py")
    print("=" * 60)


if __name__ == "__main__":
    import_md_to_excel()
